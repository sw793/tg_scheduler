"""Все операции с запланированными сообщениями канала через Pyrogram."""

import asyncio
import csv
import io
import logging
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import pytz
from openpyxl.styles import Alignment
from pyrogram import Client
from pyrogram.errors import FloodWait, ChatAdminRequired

from config import Config
from sheets import DATE_FORMATS, get_gspread_client

logger = logging.getLogger(__name__)

_MAX_FLOOD_RETRIES = 3
_DELETE_CHUNK_SIZE = 100


def _parse_naive_to_utc(dt_str: str, tz_name: str) -> datetime | None:
    """Распарсить строку даты, локализовать в tz_name и конвертировать в UTC."""
    for fmt in DATE_FORMATS:
        try:
            naive = datetime.strptime(dt_str.strip(), fmt)
            local_tz = pytz.timezone(tz_name)
            local_dt = local_tz.localize(naive)
            return local_dt.astimezone(timezone.utc).replace(tzinfo=None)
        except (ValueError, pytz.UnknownTimeZoneError):
            continue
    return None


async def _send_with_retry(
    client: Client,
    channel_id: int,
    text: str,
    schedule_date: datetime,
    row: int,
) -> str | None:
    """Отправить сообщение с повторами при FloodWait. Вернуть описание ошибки или None."""
    for attempt in range(1, _MAX_FLOOD_RETRIES + 1):
        try:
            await client.send_message(
                chat_id=channel_id,
                text=text,
                schedule_date=schedule_date,
            )
            return None
        except FloodWait as e:
            wait_sec = e.value + 2
            logger.warning(
                "Строка %d: FloodWait %d сек (попытка %d/%d)", row, wait_sec, attempt, _MAX_FLOOD_RETRIES
            )
            await asyncio.sleep(wait_sec)
        except ChatAdminRequired:
            return "Нет прав администратора в канале"
        except Exception as exc:
            return f"{type(exc).__name__}: {exc}"
    return f"FloodWait: превышено максимальное количество попыток ({_MAX_FLOOD_RETRIES})"


async def import_posts(client: Client, posts: list[dict], config: Config) -> dict:
    """Запланировать посты из списка через Pyrogram scheduled messages."""
    success: list[int] = []
    skipped: list[dict] = []
    failed: list[dict] = []

    now_utc = datetime.now(timezone.utc).replace(tzinfo=None)
    min_dt = now_utc.replace(second=now_utc.second + 300)  # now + 5 минут

    for post in posts:
        row = post["row"]
        text = post["text"]
        dt_str = post["dt_str"]

        dt_utc = _parse_naive_to_utc(dt_str, config.TIMEZONE)
        if dt_utc is None:
            skipped.append({"row": row, "reason": f"Не удалось распарсить дату: '{dt_str}'"})
            continue

        # now + 5 минут
        now_utc_fresh = datetime.now(timezone.utc).replace(tzinfo=None)
        threshold = now_utc_fresh.replace(second=0, microsecond=0)
        import datetime as dt_module
        threshold = datetime.now(timezone.utc).replace(tzinfo=None) + dt_module.timedelta(minutes=5)

        if dt_utc <= threshold:
            skipped.append({"row": row, "reason": f"Дата '{dt_str}' меньше now+5 мин"})
            continue

        error = await _send_with_retry(client, config.CHANNEL_ID, text, dt_utc, row)
        if error is None:
            success.append(row)
            logger.info("Строка %d: запланирована на %s UTC", row, dt_utc.isoformat())
        else:
            failed.append({"row": row, "error": error})
            logger.error("Строка %d: ошибка — %s", row, error)

    return {"success": success, "skipped": skipped, "failed": failed}


async def export_scheduled(client: Client, config: Config) -> list[dict]:
    """Получить все запланированные сообщения канала и вернуть список dict."""
    try:
        messages = await client.get_scheduled_messages(config.CHANNEL_ID)
    except AttributeError:
        raise RuntimeError(
            "Метод get_scheduled_messages недоступен. "
            "Убедитесь, что установлена pyrogram>=2.0.106."
        )

    local_tz = pytz.timezone(config.TIMEZONE)
    result: list[dict] = []

    for msg in messages:
        msg_id: int = msg.id
        text: str = msg.text or msg.caption or "<медиа без текста>"
        # msg.date — UTC timestamp (int или datetime)
        if isinstance(msg.date, int):
            utc_dt = datetime.fromtimestamp(msg.date, tz=timezone.utc)
        else:
            utc_dt = msg.date.replace(tzinfo=timezone.utc) if msg.date.tzinfo is None else msg.date
        local_dt = utc_dt.astimezone(local_tz)
        date_str = local_dt.strftime("%d.%m.%Y %H:%M")
        result.append({"id": msg_id, "text": text, "date": date_str})

    result.sort(key=lambda x: x["date"])
    return result


def write_export_to_sheets(posts: list[dict], config: Config) -> str:
    """Создать новый лист в Google Sheets с запланированными постами. Вернуть URL листа."""
    client = get_gspread_client(config)
    spreadsheet = client.open_by_key(config.SPREADSHEET_ID)

    sheet_title = datetime.now().strftime("export_%Y-%m-%d_%H-%M")
    worksheet = spreadsheet.add_worksheet(title=sheet_title, rows=len(posts) + 1, cols=3)

    header = [["message_id", "text", "date"]]
    rows = [[p["id"], p["text"], p["date"]] for p in posts]
    worksheet.update("A1", header + rows)

    spreadsheet_id = config.SPREADSHEET_ID
    gid = worksheet.id
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={gid}"
    return url


def write_export_to_xlsx(posts: list[dict]) -> bytes:
    """Создать .xlsx-файл в памяти с запланированными постами. Вернуть bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "scheduled_posts"

    ws.append(["message_id", "text", "date"])

    wrap = Alignment(wrap_text=True)
    for post in posts:
        ws.append([post["id"], post["text"], post["date"]])
        ws.cell(row=ws.max_row, column=2).alignment = wrap

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def delete_scheduled_posts(client: Client, ids: list[int], config: Config) -> dict:
    """Удалить запланированные сообщения по списку message_id."""
    deleted_count = 0
    not_found: list[int] = []

    for i in range(0, len(ids), _DELETE_CHUNK_SIZE):
        chunk = ids[i : i + _DELETE_CHUNK_SIZE]
        try:
            await client.delete_scheduled_messages(config.CHANNEL_ID, chunk)
            deleted_count += len(chunk)
        except Exception as exc:
            logger.error("Ошибка при удалении чанка %s: %s", chunk, exc)
            not_found.extend(chunk)

    return {"deleted": deleted_count, "not_found": not_found}


def parse_ids_from_input(data: str | bytes, fmt: str) -> list[int]:
    """Распарсить message_id из текста, xlsx-файла или csv-файла."""
    ids: set[int] = set()

    if fmt == "text":
        assert isinstance(data, str)
        for line in data.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                ids.add(int(line))
            except ValueError:
                logger.warning("parse_ids_from_input: не число: '%s'", line)

    elif fmt == "xlsx":
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val is None:
                continue
            try:
                ids.add(int(val))
            except (ValueError, TypeError):
                logger.warning("parse_ids_from_input xlsx: не число: '%s'", val)

    elif fmt == "csv":
        assert isinstance(data, bytes)
        text = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        next(reader, None)  # пропустить заголовок
        for row in reader:
            if not row:
                continue
            try:
                ids.add(int(row[0].strip()))
            except (ValueError, IndexError):
                logger.warning("parse_ids_from_input csv: не число: '%s'", row)
    else:
        raise ValueError(f"Неизвестный формат: '{fmt}'. Допустимые: text, xlsx, csv")

    return list(ids)
